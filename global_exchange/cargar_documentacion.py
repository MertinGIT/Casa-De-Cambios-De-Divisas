import os
import re
import sys
from openpyxl import load_workbook

EXCEL_PATH = "EQUIPO_09_CONTROL_DOCUMENTACION.xlsx"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Regex para clases y métodos
CLASS_PATTERN = re.compile(r"^\s*class\s+([A-Za-z0-9_]+)\(", re.MULTILINE)
METHOD_PATTERN = re.compile(r"^\s*def\s+([a-zA-Z0-9_]+)\(", re.MULTILINE)

TARGET_FILES = ["models.py", "views.py", "forms.py"]

def find_classes_and_methods(base_dir, apps_to_scan):
    docs_info = []
    for app in apps_to_scan:
        app_dir = os.path.join(base_dir, app)
        if not os.path.isdir(app_dir):
            print(f"[WARN] No se encontró la app: {app}")
            continue

        for script in TARGET_FILES:
            file_path = os.path.join(app_dir, script)
            if not os.path.exists(file_path):
                continue

            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            if script == "views.py":
                # Para vistas: detectar funciones de nivel superior
                functions = METHOD_PATTERN.findall(content)
                for func in functions:
                    docs_info.append((app, script, func))
            else:
                # Para models.py y forms.py: clases y sus métodos
                classes = CLASS_PATTERN.findall(content)
                for cls in classes:
                    docs_info.append((app, script, cls))

                    # métodos de esa clase
                    class_block_pattern = re.compile(
                        rf"class {cls}\(.*?\):([\s\S]*?)(?=^class|\Z)", re.MULTILINE
                    )
                    match = class_block_pattern.search(content)
                    if match:
                        methods = METHOD_PATTERN.findall(match.group(1))
                        for method in methods:
                            docs_info.append((app, script, method))
    return docs_info

def write_to_excel(docs_info, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    col_modulo = headers.get("Modulo", 1)
    col_script = headers.get("Script", 2)
    col_metodo = headers.get("Metodo", 3)

    row = ws.max_row + 1
    for modulo, script, metodo in docs_info:
        ws.cell(row=row, column=col_modulo).value = modulo
        ws.cell(row=row, column=col_script).value = script
        ws.cell(row=row, column=col_metodo).value = metodo
        row += 1

    wb.save(excel_path)
    print(f"[OK] Se guardaron {len(docs_info)} registros en {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python cargar_documentacion_excel.py <app1> <app2> ...")
        sys.exit(1)

    apps_to_scan = sys.argv[1:]
    print(f"Escaneando aplicaciones: {', '.join(apps_to_scan)}")

    docs_info = find_classes_and_methods(BASE_DIR, apps_to_scan)
    if docs_info:
        write_to_excel(docs_info, EXCEL_PATH)
    else:
        print("No se encontraron clases ni funciones en las apps seleccionadas.")
