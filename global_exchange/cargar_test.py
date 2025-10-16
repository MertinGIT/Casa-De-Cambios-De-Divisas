import os
import re
import sys
from openpyxl import load_workbook

EXCEL_PATH = "EQUIPO_09_CONTROL_PRUEBAS_UNITARIAS.xlsx"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_METHOD_PATTERN = re.compile(r"^\s*def (test_[a-zA-Z0-9_]+)\(", re.MULTILINE)

def find_tests(base_dir, apps_to_scan):
    tests_info = []
    for root, _, files in os.walk(base_dir):
        # saltar si no corresponde a ninguna app seleccionada
        if not any(app in root for app in apps_to_scan):
            continue
        if "tests" not in root:
            continue

        for file in files:
            if file.startswith("test_") and file.endswith(".py"):
                file_path = os.path.join(root, file)

                parts = root.split(os.sep)
                if "tests" in parts:
                    idx = parts.index("tests")
                    module = parts[idx - 1] if idx > 0 else "desconocido"
                else:
                    module = "desconocido"

                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()

                methods = TEST_METHOD_PATTERN.findall(content)
                for method in methods:
                    tests_info.append((module, file, method))
    return tests_info

def write_to_excel(tests_info, excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    col_modulo = headers.get("Modulo", 1)
    col_script = headers.get("Script", 2)
    col_metodo = headers.get("Metodo", 3)

    row = ws.max_row + 1  # continuar al final
    for modulo, script, metodo in tests_info:
        ws.cell(row=row, column=col_modulo).value = modulo
        ws.cell(row=row, column=col_script).value = script
        ws.cell(row=row, column=col_metodo).value = metodo
        row += 1

    wb.save(excel_path)
    print(f"[OK] Se guardaron {len(tests_info)} tests en {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python cargar_tests_excel.py <app1> <app2> ...")
        sys.exit(1)

    apps_to_scan = sys.argv[1:]
    print(f"Escaneando aplicaciones: {', '.join(apps_to_scan)}")

    tests_info = find_tests(BASE_DIR, apps_to_scan)
    if tests_info:
        write_to_excel(tests_info, EXCEL_PATH)
    else:
        print("No se encontraron tests en las apps seleccionadas.")
