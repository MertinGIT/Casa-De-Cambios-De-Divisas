from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.utils.dateparse import parse_date

from cliente_usuario.models import Usuario_Cliente
from clientes.models import Cliente

@login_required
def historial_usuario(request):
    """
    Muestra el historial de transacciones del usuario autenticado.

    Permite filtrar las transacciones por ID, estado, tipo, moneda y rango de fechas.
    También obtiene información de segmentación y descuentos según el cliente operativo
    asociado al usuario actual.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP con posibles parámetros GET:
            - q: texto de búsqueda.
            - campo: campo por el cual buscar (id / estado / tipo).
            - moneda o moneda_sel: abreviación de la moneda.
            - fecha_inicio y fecha_fin: fechas para filtrar el rango.

    Retorna:
        HttpResponse: Página renderizada con la lista filtrada de transacciones,
                      monedas disponibles y detalles de segmentación del usuario.
    """
    from operaciones.models import Transaccion
    from monedas.models import Moneda

    # Parámetros
    q = request.GET.get('q', '').strip()
    campo = request.GET.get('campo', '').strip()              # id / estado / tipo
    moneda = (request.GET.get('moneda') or
              request.GET.get('moneda_sel') or '').strip()    # soporta ambos nombres
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()

    # === OBTENER CLIENTE OPERATIVO ===
    clientes_asociados, cliente_operativo = obtener_clientes_usuario(request.user, request)
    
    # Filtrar transacciones por usuario Y cliente operativo
    qs = (Transaccion.objects
          .filter(usuario=request.user)
          .select_related('moneda_origen', 'moneda_destino', 'cliente')
          .order_by('-fecha'))
    
    # Filtrar por cliente operativo si existe
    if cliente_operativo:
        qs = qs.filter(cliente=cliente_operativo)

    # Filtro texto / campo
    if q and campo:
        if campo == 'id' and q.isdigit():
            qs = qs.filter(id=int(q))
        elif campo == 'estado':
            qs = qs.filter(estado__icontains=q)
        elif campo == 'tipo':
            qs = qs.filter(tipo__icontains=q)

    # Fechas (YYYY-MM-DD)
    if fecha_inicio:
        d = parse_date(fecha_inicio)
        if d:
            qs = qs.filter(fecha__date__gte=d)
    if fecha_fin:
        d = parse_date(fecha_fin)
        if d:
            qs = qs.filter(fecha__date__lte=d)

    # Filtro moneda: si aparece como origen o destino
    if moneda:
        qs = qs.filter(
            Q(moneda_origen__abreviacion=moneda) |
            Q(moneda_destino__abreviacion=moneda)
        )

    # Monedas disponibles desde transacciones del usuario (evita listar sin uso)
    monedas_set = set()
    for t in qs:
        if t.moneda_origen:
            monedas_set.add(t.moneda_origen.abreviacion)
        if t.moneda_destino:
            monedas_set.add(t.moneda_destino.abreviacion)
    monedas_disponibles = sorted(monedas_set)
    
    segmento_nombre = "Sin Segmentación"
    descuento = 0
    # === SEGMENTACIÓN SEGÚN USUARIO ===
    if cliente_operativo and cliente_operativo.segmentacion and cliente_operativo.segmentacion.estado == "activo":
        segmento_nombre = cliente_operativo.segmentacion.nombre
        if cliente_operativo.segmentacion.descuento:
            descuento = float(cliente_operativo.segmentacion.descuento)

    context = {
        'transacciones': qs,
        'q': q,
        'campo': campo,
        'moneda_sel': moneda,
        'moneda': moneda,
        'monedas_disponibles': monedas_disponibles,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        "segmento": segmento_nombre,
        "clientes_asociados": clientes_asociados,
        "cliente_operativo": cliente_operativo,
        'descuento': descuento
    }
    return render(request, 'historial_transacciones/historial_usuario.html', context)

@login_required
def detalle_transaccion(request, transaccion_id):
    """
    Muestra el detalle de una transacción específica del usuario.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP.
        transaccion_id (int): Identificador de la transacción a visualizar.

    Retorna:
        HttpResponse: Página renderizada con los datos completos de la transacción.

    Excepciones:
        Http404: Si la transacción no existe o no pertenece al usuario autenticado.
    """
    from operaciones.models import Transaccion
    transaccion = get_object_or_404(Transaccion, id=transaccion_id, usuario=request.user)
    return render(request, 'historial_transacciones/detalle_transaccion.html', {
        'transaccion': transaccion
        })

@login_required
def exportar_historial_excel(request):
    """
    Exporta el historial de transacciones del usuario en formato Excel (.xlsx).

    Genera un archivo Excel con columnas como ID, fecha, monto, monedas, tipo, tasa usada
    y estado, utilizando la librería `openpyxl`.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP.

    Retorna:
        HttpResponse: Archivo Excel para descarga directa por el navegador.

    Excepciones:
        ImportError: Si la librería `openpyxl` no está instalada.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'error': 'Instalar openpyxl'}, status=500)

    from operaciones.models import Transaccion
    qs = Transaccion.objects.filter(usuario=request.user).order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    headers = ["ID", "Fecha", "Monto", "Moneda Origen", "Moneda Destino", "Tipo", "Tasa Usada", "Estado"]
    ws.append(headers)

    for t in qs:
        ws.append([
            t.id,
            timezone.localtime(t.fecha).strftime('%d/%m/%Y %H:%M') if t.fecha else '',
            float(t.monto) if getattr(t, 'monto', None) is not None else 0,
            t.moneda_origen.abreviacion if t.moneda_origen else '',
            t.moneda_destino.abreviacion if t.moneda_destino else '',
            t.tipo,
            t.tasa_usada if getattr(t, 'tasa_usada', None) is not None else '',
            t.estado,
            
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_transacciones.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_historial_pdf(request):
    """
    Exporta el historial de transacciones del usuario en formato PDF.

    Genera un documento PDF con las transacciones ordenadas por fecha, mostrando sus
    detalles principales en formato de tabla. Utiliza la librería `reportlab`.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP.

    Retorna:
        HttpResponse: Archivo PDF para descarga directa.

    Excepciones:
        ImportError: Si la librería `reportlab` no está instalada.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
    except ImportError:
        return JsonResponse({'error': 'Instalar reportlab'}, status=500)

    from operaciones.models import Transaccion
    qs = Transaccion.objects.filter(usuario=request.user).order_by('-fecha')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=historial_transacciones.pdf'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    # Título
    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, y, "Historial de Transacciones")
    y -= 30

    # Encabezados (los mismos que en Excel)
    headers = ["ID", "Fecha", "Monto", "Moneda Origen", "Moneda Destino", "Tipo", "Tasa Usada", "Estado"]
    p.setFont("Helvetica-Bold", 9)

    # Posiciones X (ajustadas para A4 horizontalmente)
    x_positions = [40, 80, 150, 220, 300, 380, 450, 520]

    for x, h in zip(x_positions, headers):
        p.drawString(x, y, h)
    y -= 12
    p.setLineWidth(0.5)
    p.line(40, y, width - 40, y)
    y -= 10

    # Filas de datocliente_operativs
    p.setFont("Helvetica", 8)
    for t in qs:
        if y < 60:  # salto de páginacliente_operativ
            p.showPage()
            y = height - 50
            p.setFont("Helvetica-Bold", 9)
            for x, h in zip(x_positions, headers):
                p.drawString(x, y, h)
            y -= 22
            p.setFont("Helvetica", 8)

        valores = [
            str(t.id),
            timezone.localtime(t.fecha).strftime('%d/%m/%Y %H:%M') if t.fecha else '',
            str(int(t.monto)) if getattr(t, 'monto', None) is not None else '0',
            t.moneda_origen.abreviacion if t.moneda_origen else '',
            t.moneda_destino.abreviacion if t.moneda_destino else '',
            t.tipo,
            str(int(t.tasa_usada)) if getattr(t, 'tasa_usada', None) is not None else '',
            t.estado,
        ]

        for x, v in zip(x_positions, valores):
            p.drawString(x, y, str(v))

        y -= 14

    p.showPage()
    p.save()
    return response

def obtener_clientes_usuario(user,request):
    """
    Obtiene los clientes asociados a un usuario y determina cuál está operativo.

    Busca todos los clientes activos relacionados con el usuario mediante el modelo
    `Usuario_Cliente`, e identifica el cliente operativo según lo guardado en sesión.
    Si no hay uno guardado, selecciona el primero de la lista.

    Parámetros:
        user (User): Usuario autenticado.
        request (HttpRequest): Objeto de solicitud HTTP con sesión activa.

    Retorna:
        tuple:
            - clientes_asociados (list): Lista de instancias de `Cliente` asociados al usuario.
            - cliente_operativo (Cliente | None): Cliente actualmente seleccionado.
    """

     # Solo clientes activos
    usuarios_clientes = (
        Usuario_Cliente.objects
        .select_related("id_cliente__segmentacion")
        .filter(id_usuario=user, id_cliente__estado="activo")
    )
    
    clientes_asociados = [uc.id_cliente for uc in usuarios_clientes if uc.id_cliente]
    cliente_operativo = None

    # Tomar de la sesión si existe
    if request and request.session.get('cliente_operativo_id'):
        cliente_operativo = next((c for c in clientes_asociados if c.id == request.session['cliente_operativo_id']), None)

    # Si no hay sesión o ID no válido, tomar el primero
    if not cliente_operativo and clientes_asociados:
        cliente_operativo = clientes_asociados[0]

    return clientes_asociados, cliente_operativo


@login_required
def set_cliente_operativo(request):
    """
    Define el cliente operativo activo para el usuario autenticado.

    Guarda el cliente seleccionado en la sesión y devuelve en formato JSON los datos
    asociados a su segmentación (nombre y descuento) para actualizar el front-end sin
    recargar la página.

    Parámetros:
        request (HttpRequest): Objeto de solicitud HTTP con datos POST:
            - cliente_id: ID del cliente a establecer como operativo.

    Retorna:
        JsonResponse:
            - En caso de éxito: información del cliente, nombre de segmento y descuento.
            - En caso de error: mensaje indicando el motivo (cliente no encontrado o petición inválida).

    Excepciones:
        Cliente.DoesNotExist: Si el cliente no existe o está inactivo.
    """
    cliente_id = request.POST.get('cliente_id')

    if cliente_id:
        try:
            cliente = Cliente.objects.select_related("segmentacion").get(
                pk=cliente_id, estado="activo"
            )
            # Guardar en sesión
            request.session['cliente_operativo_id'] = cliente.id

            # Datos a devolver
            segmento_nombre = None
            descuento = 0
            if cliente.segmentacion and cliente.segmentacion.estado == "activo":
                segmento_nombre = cliente.segmentacion.nombre
                descuento = float(cliente.segmentacion.descuento or 0)

            return JsonResponse({
                "success": True,
                "segmento": segmento_nombre,
                "descuento": descuento,
                "cliente_nombre": cliente.nombre
            })

        except Cliente.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Cliente no encontrado"}, status=404
            )

    return JsonResponse({"success": False, "error": "Petición inválida"}, status=400)
